import models
import os, sys
import openpyxl
import pandas as pd
import tensorflow as tf
import librosa
import numpy as np
from heapq import nlargest
import random
import matplotlib.pyplot as plt
import matplotlib
import seaborn as sns
from sklearn import manifold
import itertools
from sklearn.metrics import confusion_matrix
import pickle
from sklearn.metrics import accuracy_score
import argparse

def mypredict(xtest,model):
	
	########################################################################
	y_score = model.predict(xtest)
    
    #(frame,BestOfFrame)                                                                                                                                                                                                                                                 
	class_prediction = np.argmax(y_score, axis=1)
	class_probability = np.max(y_score, axis=1)
	#print(class_prediction.shape)
	#print(class_probability)

	#choose best 3
	fincandy=[]
	#################################################################
	#round1
	onehot = {"0": 0, "1": 0, "2": 0, "3": 0, "4": 0,"5": 0,"6": 0,"7": 0,"8": 0,"9": 0,"10": 0,"11": 0,"12": 0,"13": 0,"14": 0,"15": 0,"16": 0,"17": 0,"18": 0,"19": 0}
	candy = class_prediction[class_probability >= 0.5]

	for i in range(len(candy)):
		onehot[str(candy[i])]+=1

	res = nlargest(3, onehot, key=onehot.get)

	#for 0 vote candidate
	failnum=0

	for i in range(len(res)):
		if onehot[res[i]]==0:
			failnum+=1
		else:
			fincandy.append(res[i])

	#############################################
	failnum2=0
	#round2
	if failnum>0:
		onehot2 = {"0": 0, "1": 0, "2": 0, "3": 0, "4": 0,"5": 0,"6": 0,"7": 0,"8": 0,"9": 0,"10": 0,"11": 0,"12": 0,"13": 0,"14": 0,"15": 0,"16": 0,"17": 0,"18": 0,"19": 0}
		candy2 = class_prediction[class_probability <0.5]

		for i in range(len(candy2)):
			if str(candy2[i]) not in  fincandy:
				onehot2[str(candy2[i])]+=1

		res2 = nlargest(failnum, onehot2, key=onehot2.get)
		

		for i in range(len(res2)):
			if onehot2[res2[i]]==0:
				failnum2+=1
			else:
				fincandy.append(res2[i])

	##########################################
	#round3
	if failnum2>0:
		res3=[str(i) for i in range(0, 20)]
		for i in range(len(fincandy)):
			res3.remove(fincandy[i])

		res4= random.sample(res3,failnum2)

		for i in range(failnum2):
			fincandy.append(res4[i])
	#######################################################
	return fincandy
##########################################################################

def transform(song_path):
	sr=16000
	n_mels=128
	n_fft=2048
	hop_length=512
	#######################################################
	y, sr = librosa.load(song_path, sr=sr)
	S = librosa.feature.melspectrogram(y, sr=sr, n_mels=n_mels,n_fft=n_fft,hop_length=hop_length)

	log_S = librosa.core.amplitude_to_db(S, ref=1.0)

	length = 313#157

	X = []

	slices = int(log_S.shape[1] / length)
	for j in range(slices - 1):
		X.append(log_S[:, length * j:length * (j + 1)])

	X=np.array(X)

	X = X.reshape(X.shape + (1,))

	return X
###########################################
def plot_tsne(xval,yval,mod):
	myfun=tf.keras.backend.function([mod.layers[0].input],[mod.layers[-3].output])

	X_rep = []
		
	for i in range(len(xval)):
		chec = (xval[i])[0]
		X_rep.extend(myfun([np.expand_dims(chec,axis = 0)])[0])


	X_rep = np.array(X_rep)

	tsne_model = manifold.TSNE()
	X_2d = tsne_model.fit_transform(X_rep)

	sns.set_palette("Paired", n_colors=20)
	plt.figure(figsize=(20, 20))

	sns.scatterplot(x=X_2d[:, 0], y=X_2d[:, 1],hue=np.array(yval), palette=sns.color_palette(n_colors=20))
					
	plt.savefig('tsne.png', bbox_inches="tight")
	plt.close()

##############################################################
def plot_confusion_matrix(cm):

	classes=['aerosmith','beatles','creedence_clearwater_revival','cure','dave_matthews_band',
	'depeche_mode','fleetwood_mac','garth_brooks','green_day','led_zeppelin',
	'madonna','metallica','prince','queen','radiohead','roxette',
	'steely_dan','suzanne_vega','tori_amos','u2']

	cmap=matplotlib.colormaps['Blues']
	title='Confusion matrix'
	plt.figure(figsize=(14, 14))
	
	cm = cm.astype('float') / cm.sum(axis=1)[:, np.newaxis]

	plt.imshow(cm, interpolation='nearest', cmap=cmap)

	plt.title(title)
	plt.colorbar()
	tick_marks = np.arange(len(classes))
	plt.xticks(tick_marks, classes, rotation=90)
	plt.yticks(tick_marks, classes)

	fmt='.2f'
	thresh = cm.max() / 2.
	
	for i, j in itertools.product(range(cm.shape[0]), range(cm.shape[1])):
		plt.text(j, i, format(cm[i, j], fmt),horizontalalignment="center",color="white" if cm[i, j] > thresh else "black")

	plt.tight_layout()
	plt.ylabel('True label')
	plt.xlabel('Predicted label')


	plt.savefig(title+'.png',bbox_inches="tight")
	plt.close()



if __name__ == '__main__':

	parser = argparse.ArgumentParser()
	parser.add_argument("--valid",default=None,dest='validpath',help="validation data path")
	parser.add_argument("--test",default=None,dest='testpath',help="testing data path")
	parser.add_argument("--weight",default=None,required = True,dest='weightpath',help="weight path")
	parser.add_argument("--output",default=None,dest='outputpath',help="prediction output path/filename")
	parser.add_argument("--debug",action='store_true',help="debug")

	args = parser.parse_args()
	###################################################
	#valid_mode = True
	#test_mode = False
	#debug = True
	#10s 313 frames
	slice_length=313
	#model weight path
	save_path = args.weightpath #"./myweight/myweight.ckpt"

	mymodel = models.model(1, 2, 3, 20, slice_length)
	mymodel.load_weights(save_path).expect_partial()
	############################################
	search={"0":'aerosmith', "1":'beatles', "2":'creedence_clearwater_revival', "3":'cure', "4":'dave_matthews_band',
	"5":'depeche_mode',"6":'fleetwood_mac',"7":'garth_brooks',"8":'green_day',"9":'led_zeppelin',
	"10":'madonna',"11":'metallica',"12":'prince',"13":'queen',"14":'radiohead',"15":'roxette',
	"16":'steely_dan',"17":'suzanne_vega',"18":'tori_amos',"19":'u2'}
	############################################

	if args.testpath is not None:
		
		##################################
		wb = openpyxl.Workbook()
		#wb = openpyxl.load_workbook('empty.xlsx')
		ws = wb.active
		col1="A"
		col2='B'
		col3='C'
		col4='D'
		############################################
		path = args.testpath #"./artist20_testing_data/"
		dirs = os.listdir( path )
		############################################
		count=0

		for song in dirs:
			file_title, file_type = os.path.splitext(song)
			count+=1
			cell1=col1+str(count)
			cell2=col2+str(count)
			cell3=col3+str(count)
			cell4=col4+str(count)

			spath=os.path.join(path,song)

			xx=transform(spath)

			preds=mypredict(xx,mymodel)

			ws[cell1]=str(file_title)
			ws[cell2]=search[preds[0]]
			ws[cell3]=search[preds[1]]
			ws[cell4]=search[preds[2]]

		wb.save('empty.xlsx') 
		inputExcelFile ="empty.xlsx"
		excelFile = pd.read_excel (inputExcelFile)
	    #predictionpath=os.path.join(args.outputpath, "r11943113.csv")
		excelFile.to_csv (args.outputpath, index = None, header=True)



	if args.validpath is not None:
		
		#########################################################
		path = args.validpath #"./val/"
		dirs = os.listdir( path )

		#for top1 accuracy
		y_true=[]
		y_predict=[]
		xxx=[]

		#for top3 accuracy
		y_predict2=[]


		for artist in dirs:
			#print(artist)
			albumpath=os.path.join(path,artist)
			dirs2 = os.listdir(albumpath)
			for a in dirs2:
				songspath=os.path.join(albumpath,a)
				dirs3 = os.listdir(songspath)
				for song in dirs3:
					sspath=os.path.join(songspath,song)
					xs=transform(sspath)
					frame_num=xs.shape[0]
					see=mypredict(xs,mymodel)
					pred=search[see[0]]
					y_predict.append(pred)
					y_true.append(artist)
					xxx.append(xs)
					################################
					pred2s=[]
					for k in range(len(see)):
						pred2=search[see[k]]
						pred2s.append(pred2)
					if artist in pred2s:
						y_predict2.append(artist)
					else:
						y_predict2.append(pred2s[0])

		##################################################
		myclasse=['aerosmith','beatles','creedence_clearwater_revival','cure','dave_matthews_band',
		'depeche_mode','fleetwood_mac','garth_brooks','green_day','led_zeppelin',
		'madonna','metallica','prince','queen','radiohead','roxette',
		'steely_dan','suzanne_vega','tori_amos','u2']	


		#############################################
		#with open('y_true.pkl', 'wb') as f:
		#	pickle.dump(y_true, f)
		
		#with open('y_predict.pkl', 'wb') as f:
		#	pickle.dump(y_predict, f)

		#with open('xxx.pkl', 'wb') as f:
		#	pickle.dump(xxx, f)

		#with open('y_predict2.pkl', 'wb') as f:
		#	pickle.dump(y_predict2, f)

		################################################
		#with open('y_true.pkl', 'rb') as f:
		#	y_true = pickle.load(f)

		#with open('y_predict.pkl', 'rb') as f:
		#	y_predict = pickle.load(f)

		#with open('xxx.pkl', 'rb') as f:
		#	xxx = pickle.load(f)

		#with open('y_predict2.pkl', 'rb') as f:
		#	y_predict2 = pickle.load(f)
		
		top1acc=accuracy_score(y_true, y_predict)
		print("Top 1 accuracy: {}".format(top1acc))

		top3acc=accuracy_score(y_true, y_predict2)
		print("Top 3 accuracy: {}".format(top3acc))
		#######################################

		cm=confusion_matrix(y_true, y_predict,labels=myclasse)

		plot_confusion_matrix(cm)

		plot_tsne(xxx,y_true,mymodel)

	if args.debug:
		print('debug')
		#path="./debug/test1.mp3"
		#fleetwood_mac  beatles  steely_dan

		path="./debug/test2.mp3"
		#aerosmith metallica  queen

		x=transform(path)
		see=mypredict(x,mymodel)
		pred=search[see[0]]
		pred1=search[see[1]]
		pred2=search[see[2]]
		print(pred)
		print(pred1)
		print(pred2)

